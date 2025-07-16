## this is currently a python file that I'm just learning to use 
## with Excel, and then I hope to be able to use it to find information about
## lists of books and movies I've collected

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from imdb import Cinemagoer

cine = Cinemagoer()

wb = load_workbook(r'C:\Users\Michael\Documents\Movie List.xlsx')
ws = wb.active
# print(ws['A2'].value)



# Getting movie id
#movie = cine.search_movie(ws['A2'].value)
#id = (movie[0].movieID)
#moviedata = cine.get_movie(id)

# Getting info from movie id
#for director in moviedata['directors']:
#    print(director['name'])
#    ws['B2'].value = director['name']
#print(moviedata['title'])
#print(moviedata['genres'])
# print(moviedata['runtime'])
# print(moviedata['rating'])
# print(moviedata['year'])

for row in range(2,68):
    for col in range(1, 8):
        if col == 1:    
            char = get_column_letter(col)
            target = (ws[char + str(row)].value)
            movie = cine.search_movie(target)
            id = (movie[0].movieID)
            moviedata = cine.get_movie(id)
            print(target)
        elif col == 2:
            char = get_column_letter(col)
            try:
                if moviedata['directors']:
                    for director in moviedata['directors']:
                        ws[char + str(row)].value = director['name']
                        print(ws[char + str(row)].value)
            except KeyError:
                next
        elif col == 3:
            char = get_column_letter(col)
            try:
                if moviedata['title']:
                        ws[char + str(row)].value = moviedata['title']
                        print(ws[char + str(row)].value)
            except KeyError:
                next
        elif col == 4:
            char = get_column_letter(col)
            try:
                if moviedata['genres']:
                        ws[char + str(row)].value = str(moviedata['genres'])
                        print(ws[char + str(row)].value)
            except KeyError:
                next
        elif col == 5:
            char = get_column_letter(col)
            try:
                if moviedata['runtime']:
                        ws[char + str(row)].value = str(moviedata['runtime'])
                        print(ws[char + str(row)].value)
            except KeyError:
                next
        elif col == 6:
            char = get_column_letter(col)
            try:
                if moviedata['rating']:
                        ws[char + str(row)].value = moviedata['rating']
                        print(ws[char + str(row)].value)
            except KeyError:
                next
        elif col == 7:
            char = get_column_letter(col)
            try:
                if moviedata['year']:
                        ws[char + str(row)].value = moviedata['year']
                        print(ws[char + str(row)].value)
            except KeyError:
                next

wb.save(r'C:\Users\Michael\Documents\Movie List.xlsx')