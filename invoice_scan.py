#! python3
# to scan invoices
# this was written by GPT so there's going to be some trouble shooting involved


## TODO: I'm not finding the invoice information / invoice number on sheet metal supply or pfs invoices, looks like there is 
# legible invoice info on Stevenson
# Hinckley is scannable

# PFS lacks a date pattern, line 239, will also need some help figuring out where to scan for it
import cv2
import pypdf
import pytesseract
import pdf2image
import os, re, datetime
import openpyxl
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook
from datetime import date

from tkinter import Tk, simpledialog, messagebox

# You must install tesseract on your machine and update the path below
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Function to perform OCR
def ocr_core(image):
    """
    This function will handle the OCR processing of images.
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Get the height and width of the image
    height, width = gray.shape
    
    # Define the bottom right quadrant
    x = 2 * width // 4
    y = 3 * height // 4
    w = width // 2
    h = height // 4

    # Crop the region of interest for the invoice number
    inv_roi = gray[0:height//2, width//2:width]
    inv_text = pytesseract.image_to_string(inv_roi, config='--psm 6')

    # Crop the region of interest for the amount
    amt_roi = gray[y:y+h, x:x+w]
    
    amt_text = pytesseract.image_to_string(amt_roi, config='--psm 6')

    text = pytesseract.image_to_string(gray)

    return text, inv_text, amt_text

def extract_text_from_pdf(file_path):
    pdf_file_obj = open(file_path, 'rb')
    pdf_reader = pypdf.PdfReader(pdf_file_obj)
    num_pages = len(pdf_reader.pages)
    text = ""
    for page in range(num_pages):
        #page_obj = pdf_reader.getPage(page)
        page_obj = pdf_reader.pages[page]
        text += page_obj.extract_text()
    pdf_file_obj.close()
    return text

def is_file_in_folder(filename, folder_path):
    file_path = os.path.join(folder_path, filename)
    return os.path.isfile(file_path)

# working on turning this into oop because this is a little convoluted
class Vendor:
    def __init__(self, name, inv_patterns, inv_location, 
                 page_pattern, total_patterns, total_location, date_pattern, date_location):
        self.name = name
        self.inv_patterns = inv_patterns
        self.inv_location = inv_location
        self.page_pattern = page_pattern
        self.total_patterns = total_patterns
        self.total_location = total_location
        self.date_pattern = date_pattern
        self.date_location = date_location


               



file_name = f'{date.today()}_invoices.xlsx'

folder_path = r'C:\Users\Michael\Desktop\python-work\Invoices'

if is_file_in_folder(file_name, folder_path):
    print('success')
    wb = load_workbook(os.path.join(folder_path, file_name))
    ws = wb.active
else:
    wb = Workbook()

    ws = wb.active

    ws['A1'] = 'Vendor'
    ws['B1'] = 'Date'
    ws['C1'] = 'Number'
    ws['D1'] = 'Amount Due'

invoice_counter = len(ws['C'])
unique_invoices = []

for cell in ws['C']:
    unique_invoices.append(cell.value)

## Loops through the files in a chosen folder, extracts the relevant text from them
## and puts them in an excel file
for filename in os.listdir(folder_path):
    if filename.endswith(".pdf"): # add any file type you want to process
        print(os.path.join(folder_path, filename))
        # Replace 'path_to_your_pdf' with the path to the PDF file you want to process
                
        pages = convert_from_path(os.path.join(folder_path, filename), dpi=300)

        # Perform OCR on each image
        invoice_no = None

        for i, page in enumerate(pages):
            vendor = None
            match_counter = 0
            page.save('out.jpg', 'JPEG')
            img = cv2.imread('out.jpg')
            # print(f"Content on {ocr_core(img)}")
            content, inv_text, amt_text = ocr_core(img)
            print(f'content:{content}')
            print(f'amt_text:{amt_text}')
            

            # Finds the vendor
            if content.startswith("GEN") or content.startswith("GEM"):
                print("Vendor: GEMCO")
                
                vendor = "Gemco"
            elif content.startswith("INVOICE\nPro Fastening Systems") or content.startswith("Pro Fastening Systems") or content.startswith("INVOICE\n\nPro Fastening Systems"):
                print("Vendor: Pro Fastening Systems")
                ws[f'A{invoice_counter+1}'] = 'Pro Fastening Systems'
                vendor = "Pro Fastening Systems"
            elif content.startswith("Sheet Metal Supply"):
                print("Vendor: Sheet Metal Supply")
                vendor = "Sheet Metal Supply"
            elif content.startswith("Stevenson Crane Service"):
                print("Vendor: Stevenson Crane")
                vendor = "Stevenson Crane"
            elif "BEACON" in content:
                vendor = "Beacon"
                print(f"Vendor: {vendor}")
            elif "Lumber & Supply" in content:
                vendor = "Totem"
                print(f"Vendor: {vendor}")
            
            if vendor == "Gemco":

                Gemco = Vendor('Gemco', 
                    [r'\n\n(\d{7,10}-\d{1,3})', r'(\d{7,10}-\d{1,3})'], 
                    content, 
                    r"page\s(\d+)\sof\s([2-9])",
                    r'balance \$([\d,]*\.\d{2}|\d+)', 
                        amt_text,
                        r'invoice date[:\s]*([01]?\d/[0123]?\d/\d{2})', 
                        content)
                ## Finds the invoice number and checks for uniqueness
                inv_pattern = r'\n\n(\d{7,10}-\d{1,3})'  # Matches 7 to 10 digits, a hyphen, then 1 to 3 digits
                inv_pattern2 = r'(\d{7,10}-\d{1,3})'   # Same as above but with an optional hyphen
                ## Page pattern is to check whether there is text that says
                # Page x of x because if it does then we need to 
                # process things differently than if it's a single page document
                page_pattern = r"page\s(\d+)\sof\s([2-9])"

                inv_match = re.search(inv_pattern, content, re.IGNORECASE)
                inv_match2 = re.search(inv_pattern2, content, re.IGNORECASE)
                
                if inv_match:
                    match_counter = 1   
                elif inv_match2:
                    match_counter = 2
                page_match = re.search(page_pattern, content, re.IGNORECASE)
                # print(inv_match)
                if match_counter:
                    if match_counter == 1:
                        invoice_no = inv_match.group(1)
                    else:
                        invoice_no = inv_match2.group(1)

                    print(f"Found 'invoice number': {invoice_no}")
                    if invoice_no and invoice_no not in unique_invoices:
                        invoice_counter += 1
                        unique_invoices.append(invoice_no)
                        ws[f'A{invoice_counter}'] = 'Gemco'
                        ws[f'C{invoice_counter}'] = f'{invoice_no}'
                        
                        # # Finds the vendor
                        # if content.startswith("GEN") or content.startswith("GEM"):
                        #     print("Vendor: GEMCO")
                        #     ws[f'A{invoice_counter}'] = 'Gemco'
                        # elif content.startswith("Pro Fastening"):
                        #     print("Vendor: Pro Fastening Systems")
                        #     ws[f'A{invoice_counter}'] = 'Pro Fastening Systems'

                        # Finds the date of the invoice    
                        date_match = re.search(r'invoice date[:\s]*([01]?\d/[0123]?\d/\d{2})', content, re.IGNORECASE)
                        if date_match:
                            # If a match was found, 'group(1)' contains the first parenthesized subgroup - the date.
                            invoice_date = date_match.group(1)
                            print(f"Found 'invoice date': {invoice_date}")
                            ws[f'B{invoice_counter}'] = f'{invoice_date}'
                        else:
                            print("No 'invoice date' found")

                        ## Finds the invoice amount due
                        total_match = re.search(r'balance \$([\d,]*\.\d{2}|\d+)', amt_text, re.IGNORECASE)
                        if total_match:
                            # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                            balance_amount = total_match.group(1)
                            print(f"Found 'balance amount': ${balance_amount}")
                            ws[f'D{invoice_counter}'].value = balance_amount
                        else:
                            print("No 'balance amount' found")

                if page_match:
                    total_match = re.search(r'balance \$([\d,]+\.\d{2})', content, re.IGNORECASE)
                    if total_match:
                        # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                        balance_amount = total_match.group(1)
                        print(f"Found 'balance amount': ${balance_amount}")
                        ws[f'D{invoice_counter}'] = balance_amount
                else:
                    print("No 'unique invoice number' found")
            elif vendor == "Pro Fastening Systems":
                PFS = Vendor('Pro Fastening Systems', 
                    [r'(\d{2}-\d{2}-\d{2})\s*\|?\s*(\d{7}-?\d{0,2})'],
                    inv_text, 
                    r"page\s(\d+)\sof\s([2-9])",
                        [r'THIS AMOUNT \$(\d?\s?(?:[:,])?[\d,]+\s*\.{0,1}\s*\d{2})', r'PLERSE PAT \$([\d,]+\.\d{2})'],
                        amt_text,
                        r'invoice date[:\s]*([01]?\d/[0123]?\d/\d{2})', 
                        content)
                ## Finds the invoice number and checks for uniqueness
                inv_pattern = r'(\d{2}-\d{2}-\d{2})\s*\|?\s*(\d{7}-?\d{0,2})'

                # inv_pattern2 = r'office (\d+-?\d*)'
                # page_pattern = r"page\s(\d+)\sof\s([2-9])"

                inv_match = re.search(inv_pattern, inv_text, re.IGNORECASE)

                if inv_match:
                    match_counter = 1
                # inv_match2 = re.search(inv_pattern2, inv_text, re.IGNORECASE)
                # if inv_match2:
                #     match_counter = 2
                # page_match = re.search(page_pattern, inv_text, re.IGNORECASE)
                print(inv_match)
                if match_counter:
                    if match_counter == 1:
                        invoice_no = inv_match.group(2)
                        invoice_date = inv_match.group(1)
                        print(f"Found 'invoice date': {invoice_date}")
                        

                    print(f"Found 'invoice number': {invoice_no}")
                    if invoice_no not in unique_invoices:
                        invoice_counter += 1
                        unique_invoices.append(invoice_no)
                        ws[f'A{invoice_counter}'] = 'Pro Fastening Systems'
                        ws[f'B{invoice_counter}'] = f'{invoice_date}'
                        ws[f'C{invoice_counter}'] = f'{invoice_no}'

                        ## Finds the invoice amount due
                        # total_match = re.search(r'THIS AMOUNT \$(\d?\s?(?:[:,])?[\d,]+\.\d{2})', amt_text, re.IGNORECASE)
                        total_match = re.search(r'THIS AMOUNT \$(\d?\s?(?:[:,])?[\d,]+\s*\.{0,1}\s*\d{2})', amt_text, re.IGNORECASE)
                        

                        total_match2 = re.search(r'PLERSE PAT \$([\d,]+\.\d{2})', amt_text, re.IGNORECASE)
                        if total_match:
                            # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                            balance_amount = total_match.group(1).replace(" ", ".")
                            balance_amount =  balance_amount.replace(' :', ',')
                            balance_amount = balance_amount.replace(":", ",")
                            print(f"Found 'balance amount': ${balance_amount}")
                            ws[f'D{invoice_counter}'].value = balance_amount
                        elif total_match2:
                            balance_amount =  total_match2.group(1).replace(' :', ',')
                            balance_amount = balance_amount.replace(":", ",")
                            print(f"Found 'balance amount': ${balance_amount}")
                            ws[f'D{invoice_counter}'].value = balance_amount
                        else:
                            print("No 'balance amount' found")

                else:
                    print("No 'unique invoice number' found")
            elif vendor == "Beacon":
                ## TODO: still gotta work on finding the total
                ## Finds the invoice number and checks for uniqueness
                inv_number_pattern = r'([A-Z]+\d+)'
                inv_date_pattern = r'INVOICE DATE (\d{2}/\d{2}/\d{2})'


                # inv_pattern2 = r'office (\d+-?\d*)'
                # page_pattern = r"page\s(\d+)\sof\s([2-9])"

                inv_number_match = re.search(inv_number_pattern, inv_text)
                inv_date_match = re.search(inv_date_pattern, inv_text)
                if inv_number_match:
                    invoice_number = inv_number_match.group(1)
                    print("Invoice Number:", invoice_number)
                else:
                    print("Invoice number not found in the text")
                    
                if inv_date_match:
                    invoice_date = inv_date_match.group(1)
                    print("Invoice Date:", invoice_date)
                else:
                    print("Invoice date not found in the text")
                if invoice_number not in unique_invoices:
                    invoice_counter += 1
                    unique_invoices.append(invoice_no)
                    ws[f'A{invoice_counter}'] = 'Beacon'
                    ws[f'B{invoice_counter}'] = f'{invoice_date}'
                    ws[f'C{invoice_counter}'] = f'{invoice_number}'

                    ## Finds the invoice amount due
                    # decimal_pattern = r'(?:\d*\.\d{2}\s*){5}'
                    decimal_pattern = r'(?:(?:\d*\.(?:\d{2}|00)|00)\s*){5}'

                    total_match = re.search(decimal_pattern, content)
                    print(total_match)
                    # line_of_text = "1597.50 127.80 .00 230.78 1956.08"
                    # total_match = re.search(decimal_pattern, line_of_text)

                    # Define a pattern to match five decimal numbers separated by spaces

                    # Find all matches of the decimal pattern in the line
                    matches = re.findall(decimal_pattern, content)
                    print(total_match)
                   
                    if total_match:
                        # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                        
                        balance_amount =  total_match.group(0).replace(' :', ',')
                        balance_amount = balance_amount.replace(":", ",")
                        balance_amount = balance_amount.split()[-1]
                        print(f"Found 'balance amount': ${balance_amount}")
                        ws[f'D{invoice_counter}'].value = balance_amount
                    elif total_match2:
                        balance_amount =  total_match2.group(1).replace(' :', ',')
                        balance_amount = balance_amount.replace(":", ",")
                        print(f"Found 'balance amount': ${balance_amount}")
                        ws[f'D{invoice_counter}'].value = balance_amount
                    else:
                        print("No 'balance amount' found")

                else:
                    print("No 'unique invoice number' found")
            elif vendor == "Sheet Metal Supply":
                ## Finds the invoice number and checks for uniqueness
                inv_pattern = r'DATE INVOICE #\n([01]?\d/[0123]?\d/\d{4}) (\d{6}|\d+-\d{2})'
                # inv_pattern2 = r'office (\d+-?\d*)'
                # page_pattern = r"page\s(\d+)\sof\s([2-9])"

                inv_match = re.search(inv_pattern, inv_text, re.IGNORECASE)

                if inv_match:
                    match_counter = 1
                # inv_match2 = re.search(inv_pattern2, inv_text, re.IGNORECASE)
                # if inv_match2:
                #     match_counter = 2
                # page_match = re.search(page_pattern, inv_text, re.IGNORECASE)
                print(inv_match)
                if match_counter:
                    if match_counter == 1:
                        invoice_no = inv_match.group(2)
                        invoice_date = inv_match.group(1)
                        print(f"Found 'invoice date': {invoice_date}")
                        

                    print(f"Found 'invoice number': {invoice_no}")
                    if invoice_no not in unique_invoices:
                        invoice_counter += 1
                        unique_invoices.append(invoice_no)
                        ws[f'A{invoice_counter}'] = 'Sheet Metal Supply'
                        ws[f'B{invoice_counter}'] = f'{invoice_date}'
                        ws[f'C{invoice_counter}'] = f'{invoice_no}'

                        ## Finds the invoice amount due
                        total_match = re.search(r'Balance Due \$([\d,]*\.\d{2}|\d+)', amt_text, re.IGNORECASE)
                        if total_match:
                            # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                            balance_amount = total_match.group(1)
                            print(f"Found 'balance amount': ${balance_amount}")
                            ws[f'D{invoice_counter}'].value = balance_amount
                        else:
                            print("No 'balance amount' found")

                else:
                    print("No 'unique invoice number' found")
            elif vendor == "Stevenson Crane":
                ## Finds the invoice number and checks for uniqueness
                inv_pattern = r'Invoice (\d+)\nInvoice Date: (Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday), ([a-zA-Z]+) (\d{1,2}), (\d{4})'

                # inv_pattern2 = r'office (\d+-?\d*)'
                # page_pattern = r"page\s(\d+)\sof\s([2-9])"

                inv_match = re.search(inv_pattern, inv_text, re.IGNORECASE)

                if inv_match:
                    match_counter = 1
                # inv_match2 = re.search(inv_pattern2, inv_text, re.IGNORECASE)
                # if inv_match2:
                #     match_counter = 2
                # page_match = re.search(page_pattern, inv_text, re.IGNORECASE)
                print(inv_match)
                if match_counter:
                    if match_counter == 1:
                        invoice_no = inv_match.group(1)
                        month_name = inv_match.group(3)
                        day = inv_match.group(4)
                        year = inv_match.group(5)[-2:]  # Get the last two digits of the year

                        # Mapping of month names to their corresponding numbers
                        month_mapping = {
                            'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
                            'July': '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'
                        }

                        month_number = month_mapping[month_name]

                        # Combine into a single mm/DD/yy string
                        invoice_date = f"{month_number}/{day.zfill(2)}/{year}"
                        print(f"Found 'invoice date': {invoice_date}")
                        

                    print(f"Found 'invoice number': {invoice_no}")
                    if invoice_no not in unique_invoices:
                        invoice_counter += 1
                        unique_invoices.append(invoice_no)
                        ws[f'A{invoice_counter}'] = 'Stevenson Crane'
                        ws[f'B{invoice_counter}'] = f'{invoice_date}'
                        ws[f'C{invoice_counter}'] = f'{invoice_no}'

                        ## Finds the invoice amount due
                        # total_match = re.search(r'Total Invoice:\s*\$([\d,]+\.\d{2})', amt_text, re.IGNORECASE)
                        total_match = re.search(r'Total Invoice:\s*\$([\d,]*\s*(?:\.\d{2})?|\d+\s*)', amt_text, re.IGNORECASE)
                        total_match2 = re.search(r'Total Invoice:\s*\$([\d,]*\s*(?:\.\d{2})?|\d+\s*)', content, re.IGNORECASE)


                        if total_match:
                            # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                            balance_amount = total_match.group(1)
                            print(f"Found 'balance amount': ${balance_amount}")
                            ws[f'D{invoice_counter}'].value = balance_amount
                        elif total_match2:
                            # If a match was found, 'group(1)' contains the first parenthesized subgroup - the balance amount.
                            balance_amount = total_match2.group(1)
                            print(f"Found 'balance amount': ${balance_amount}")
                            ws[f'D{invoice_counter}'].value = balance_amount
                        else:
                            print("No 'balance amount' found")

                else:
                    print("No 'unique invoice number' found")
            
    else:
        continue

for i in range(2,len(ws['D'])+1):
    print(ws[f'D{i}'].value)
    if ws[f'D{i}'].value is None:
        missing_invoice = simpledialog.askstring('Missing Total', f"What is the balance due for invoice {ws[f'C{i}'].value}")
        ws[f'D{i}'].value = missing_invoice

wb.save(os.path.join(folder_path, file_name))
