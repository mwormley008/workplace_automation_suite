# This loops through the newly created dictionary with the categories as keys
# You start this in QB with the vendor, ref number, and amount due already put in
# and your cursor should start in the account bubbleimport pyautogui, openpyxl, datetime, calendar, os, sys, re

import subprocess
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

from tkinter import Tk, simpledialog, messagebox
from tkinter.filedialog import askopenfilename
from pyautogui import press, write, hotkey

from time import sleep



initial_dir=r"C:\Users\Michael\Documents"
folder_path = initial_dir 

workbook_path = askopenfilename(initialdir=initial_dir)

excel_sheet = workbook_path
invoices_for_input = load_workbook(excel_sheet, data_only=True)

sleep(1)



def quick_windows():
    windows = gw.getAllWindows()

    qb_window = None

    for window in windows:
        if "QuickBooks Desktop" in window.title:
            qb_window = window

    return qb_window 
ws = invoices_for_input['Sheet1']
# This gets the bottom row. While I was testing for some reason it was going one row over
# so i just said length minus one, but maybe that won't be true in the future
# bottom_row = len(ws['A'])-1
bottom_row = 0
for row, cell in enumerate(ws['A']):
    if cell.value is not None and cell.value != '':
        bottom_row = row

# Since row indexing starts from 0, you might want to add 1 to get the actual row number
bottom_row += 1

print(bottom_row)
print(bottom_row)
Amount = False

row_cells = ws[bottom_row]
bill_total = 0
categories_to_pay = {} 

for cell in row_cells:
    if bill_total and bill_total == cell.value:
        break
    else:
        bill_total += cell.value
    if cell.value:
        
        column_letter_entry = {cell.column_letter:round(cell.value, 2)}
        categories_to_pay.update(column_letter_entry) 

    # print(cell.value)
    # print(categories_to_pay)

categories = {'A': '871', 'B': '887', 'C': '872', 'D': '702', 'E': '908', 'F': '917', 'G': '889', 'H': '914', 'I': '860', 'J': '878', 'K': '893', 'L': '268', 'M': '886', 'N': '880', 'O': '883', 'P': '854', 'Q': '869', 'R': '255', 'S': '905'}

values_with_categories_as_keys = {categories[key]: value for key, value in categories_to_pay.items()}
print(values_with_categories_as_keys)

# Alright, this tsarts with the bill window open in QB

qb_window = quick_windows()
sleep(1)

if qb_window:
    qb_window.activate()
else:
    print("QuickBooks window not found.")   

sleep(1)


# This loops through the newly created dictionary with the categories as keys
# You start this in QB with the vendor, ref number, and amount due already put in
# and your cursor should start in the account bubble

for category, value in values_with_categories_as_keys.items():
    sleep(1)
    write(category)
    sleep(.5)
    press('tab')
    sleep(.5)
    write(str(value))
    sleep(1)
    press('tab', presses=3)