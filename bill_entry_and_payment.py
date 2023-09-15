# This is a program to use an excel spreadsheet to enter your billing information and then
# Run the rest of the process by itself, including writing the checks and
# Creating envelopes as necessary

# Currently starts from tphe enter bill screen

# Maybe what i should do is prompt to see if we want to use the default mail settings

import pyautogui, openpyxl, datetime, calendar, os, sys, re
import subprocess
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

from tkinter import Tk, simpledialog, messagebox
from tkinter.filedialog import askopenfilename
from pyautogui import press, write, hotkey
from time import sleep
import win32com.client


windows = gw.getAllWindows()

qb_window = None

for window in windows:
    if "QuickBooks" in window.title:
        qb_window = window
        break

workbook_path = workbook_path = r'C:\Users\Michael\Desktop\python-work\Vendors.xlsx'

wb = openpyxl.load_workbook(workbook_path)

ws = wb['Sheet1']

bill_status = messagebox.askyesno("Bill Entry", "Do you need to enter new bills from the sheet to quickbooks?")
default_mail_status = messagebox.askyesno("Default mail", "Do you want to use the default envelope settings for each vendor?")
print(bill_status)

first_row_skipped = False
# List to store row numbers with a cell value
rows_with_value = []

check_in_excel = messagebox.askyesno("Check number", f"Is the first check number {ws['I1'].value}")
if check_in_excel:
    check_numbers = ws['I1'].value
else:
    check_numbers = simpledialog.askinteger("Check numbers", "What is the first check number?")

qb_window.activate()
sleep(1)
# Replace 'A' with the column letter you want to check (e.g., 'B', 'C', 'D', etc.)

column_to_check = 'B'
for cell in ws[column_to_check]:
    # Skip the first row (header row)
    if not first_row_skipped:
        first_row_skipped = True
        continue

    if cell.value is not None:
        print(f"Cell {cell.coordinate} contains text: {cell.value}")
        rows_with_value.append(cell.coordinate[1:])

print(rows_with_value)

# # Write the bills to QB
qb_window.activate()
sleep(1)

if bill_status:
    for j in rows_with_value:
        sleep(1)
        if ws['A'+j].value == "Nicor Gas":
            print("nicor")
            hotkey('ctrl', 't')
            sleep(1)
            press('n')
            sleep(1)
            hotkey('alt', 't')
            press('enter')
            sleep(1)
        pyautogui.write(ws['A'+j].value)
        sleep(1)
        pyautogui.press('tab', presses=2)
        sleep(1)
        if ws['C'+j].value is not None:
            pyautogui.write(str(ws['C'+j].value))
            sleep(.5)
        pyautogui.press('tab')
        sleep(.5)
        pyautogui.write(str(ws['D'+j].value))
        sleep(.5)
        hotkey('alt', 's')
        sleep(2)

press('alt')
sleep(.5)
press('o')
# sleep(.5)
press('p')
sleep(1)


# Pay the bills

bills_counter = 0
for j in rows_with_value:
    press('tab')
    sleep(1)
    sleep(1)
    pyautogui.write(ws['A'+j].value)
    sleep(2)
    press('tab')
    sleep(2)
    press('tab')
    sleep(2)
    press('space')
    sleep(1)
    press('p')
    sleep(2)
    bills_counter += 1
    if bills_counter == len(rows_with_value):
        press('tab')
        sleep(2)
        press('enter')
        sleep(2)
    else:
        press('enter')
        sleep(3)

sleep(1)
press('tab')
sleep(.2)
write(str(check_numbers))
sleep(1)
press('enter')

print_ready = messagebox.askyesno("Ready?", "Have your checks printed?")
# pyautogui.write(ws['A'+'2'].value)
# Print any checks needed
if default_mail_status == True:
    mail_col = 'G'
else:
    mail_col = 'F'

for row in rows_with_value:
    if ws[mail_col+row]:
        a_value = ws['A' + str(row)].value
        # Path to your Word document
        print(a_value)
    
        file_path = f"C:\\Users\\Michael\\Desktop\\CEnvelopes\\{a_value}.docx"

        # Start an instance of Word
        word = win32com.client.Dispatch('Word.Application')

        # Open the document

        doc = word.Documents.Open(file_path)

        # Print the document to the default printer
        doc.PrintOut()

        # Close the documentC
        doc.Close()

        # Quit Word
        word.Quit()
        sleep(2)