import pyautogui, openpyxl, datetime, calendar, os, sys, re
import subprocess, time
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

from tkinter import Tk, simpledialog, messagebox
from tkinter.filedialog import askopenfilename, Frame, Button
from time import sleep
from pyautogui import press, write, hotkey



file_name = f'{date.today()}_invoices.xlsx'

folder_path = r'C:\Users\Michael\Desktop\python-work\Invoices'

invoices_for_input = load_workbook(os.path.join(folder_path, file_name))

def quick_windows():
    windows = gw.getAllWindows()

    qb_window = None

    for window in windows:
        if "QuickBooks Desktop" in window.title:
            qb_window = window

    return qb_window        

ws = invoices_for_input['Sheet']

first_row_skipped = False

# Alright, this starts with the bill window open in QB

qb_window = quick_windows()
sleep(1)

if qb_window:
    qb_window.activate()
else:
    print("QuickBooks window not found.")   

sleep(3)

for i in range(2,len(ws['A'])+1):
    print(ws[f'A{i}'].value)
    print(ws[f'B{i}'].value)
    print(ws[f'C{i}'].value)
    print(ws[f'D{i}'].value)
    sleep(1)
    pyautogui.write(ws[f'A{i}'].value)
    sleep(1)
    pyautogui.press('tab')
    sleep(1)
    pyautogui.write(ws[f'B{i}'].value)
    sleep(1)
    pyautogui.press('tab')
    sleep(1)
    if ws[f'C{i}'].value is not None:
        pyautogui.write(ws[f'C{i}'].value)
        sleep(.5)
    pyautogui.press('tab')
    sleep(.5)
    pyautogui.write(str(ws[f'D{i}'].value))
    sleep(.5)
    hotkey('alt', 's')
    sleep(2)

print('Entered!')

