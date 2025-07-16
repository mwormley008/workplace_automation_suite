# This is a program to use an excel spreadsheet to enter your billing information and then
# Run the rest of the process by itself, including writing the checks and
# Creating envelopes as necessary

# Currently starts from the enter bill screen


import pyautogui, openpyxl, datetime, calendar, os, sys, re
import subprocess
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

from tkinter import Tk, simpledialog, messagebox
from tkinter.filedialog import askopenfilename
from pyautogui import press, write, hotkey
from time import sleep

windows = gw.getAllWindows()

qb_window = None

for window in windows:
    if "QuickBooks" in window.title:
        qb_window = window
        break

workbook_path = workbook_path = r'C:\Users\Michael\Desktop\python-work\cote.xlsx'

wb = openpyxl.load_workbook(workbook_path)

ws = wb['Sheet1']

first_row_skipped = False

# List to store row numbers with a cell value

column_values = {"A":"PO", 
                 "B":"Cost", 
                 "C":"Lakes of Boulder Ridge", 
                 "D":"Ponds of Stoney Creek",
                 "E": "Fox Pointe",
                 "F": "Keyton Farms",
}
jobs = []

qb_window.activate()
sleep(1)
# Replace 'A' with the column letter you want to check (e.g., 'B', 'C', 'D', etc.)

for row in ws.iter_rows(min_col=3, max_col=6):
    # Skip the first row (header row)
    if not first_row_skipped:
        first_row_skipped = True
        continue
    for cell in row:
        if cell.value is not None:
            print(f"Cell {cell.coordinate} contains text: {cell.value}")
            jobs.append(cell.column_letter)
            print(jobs)


# Write the bills to QB
qb_window.activate()
sleep(1)

# From the memorized transaction list, this creates a cote inv
for i, j in enumerate(jobs):
    hotkey('alt', 't')
    print(j)
    sleep(1)
    hotkey('alt','j')
    sleep(1)
    # Writes the job name
    write(f"Tim Cote Inc.:{column_values[jobs[i]]}")
    sleep(2)
    # all of these tabs move us to the PO number
    press('tab')
    sleep(1)
    pyautogui.press('tab', presses=2)
    sleep(2)
    pyautogui.press('tab', presses=4)
    sleep(2)
    # Writes the PO number
    write(str(ws['A'+str(i+2)].value))
    sleep(1)
    pyautogui.press('tab', presses=3)
    sleep(.5)
    pyautogui.write(str(ws['B'+str(i+2)].value))
    sleep(.5)
    hotkey('alt', 'a')
    sleep(3)
    


#print invoices

hotkey('alt','f')
sleep(1)
press('f')
sleep(1)
press('i')
sleep(2)
press('enter')