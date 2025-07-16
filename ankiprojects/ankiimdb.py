from imdb import IMDb
from openpyxl import load_workbook

# Initialize IMDbPY
ia = IMDb()

file_path = r"C:\Users\Michael\Desktop\python-work\ankiprojects\movietrivia.xlsx"

wb = load_workbook(file_path)
sheet = wb.active

# Define column indices (adjust based on your Excel structure)
title_column = 1  # Column for movie titles
director_column = 2  # Column where directors will be written
actors_column = 3 # Column where actors will be written
plot_column = 4  # Column where plots will be written

# Iterate over rows, starting from the second row (assuming the first row is headers)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=1):  # Only read titles
    title_cell = row[0]  # Movie title cell
    title = title_cell.value
    if not title:
        continue

    # Search movie by title
    search_results = ia.search_movie(title)
    if not search_results:
        print(f"Movie not found: {title}")
        sheet.cell(row=title_cell.row, column=director_column, value="Not Found")
        sheet.cell(row=title_cell.row, column=actors_column, value="Not Found")
        sheet.cell(row=title_cell.row, column=plot_column, value="Not Found")
        continue

    # Get the first result and fetch full movie details
    movie = ia.get_movie(search_results[0].movieID)

    # Extract directors and plot outline
    canonical_title = movie.get('title', movie.get('title', 'Unknown Title'))
    directors = [person['name'] for person in movie.get('director', [])]
    actors = [person['name'] for person in movie.get('cast', [])[:5]]  # Limit to top 5 actors
    plot = movie.get('plot outline', 'Not Available')

    # Overwrite the title column with the canonical name
    sheet.cell(row=title_cell.row, column=title_column, value=canonical_title)

    # Update Excel
    sheet.cell(row=title_cell.row, column=director_column, value=", ".join(directors))
    sheet.cell(row=title_cell.row, column=actors_column, value=", ".join(actors))
    sheet.cell(row=title_cell.row, column=plot_column, value=plot)

    print(f"Processed: {canonical_title}")

# Save the updated Excel file
wb.save(file_path)
print("Excel file updated successfully!")