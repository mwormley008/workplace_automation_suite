# This is a program to use an excel spreadsheet to enter your billing information and then
# Run the rest of the process by itself, including writing the checks and
# Creating envelopes as necessary

# Currently starts from tphe enter bill screen

# Maybe what i should do is prompt to see if we want to use the default mail settings

import pyautogui, openpyxl, datetime, calendar, os, sys, re
import subprocess
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

from tkinter import Tk, simpledialog, messagebox
from tkinter.filedialog import askopenfilename
from pyautogui import press, write, hotkey
from time import sleep
import win32com.client
import calendar
from datetime import date

def open_pay_bill_menu():
    press('alt')
    sleep(.5)
    press('o')
    # sleep(.5)
    press('p')
    sleep(1)
def set_payment_date():
    #Sets the date to be used in the format 10/31/2023
    today = date.today()
    print(today)
    res = calendar.monthrange(today.year, today.month)[1]
    payment_date = f"{today.month:02d}/{res:02d}/{today.year}"
    return payment_date

print("Beginning monthly_bill_payment.py")

windows = gw.getAllWindows()

qb_window = None

for window in windows:
    if "QuickBooks" in window.title:
        qb_window = window
        break

workbook_path = workbook_path = r'C:\Users\Michael\Desktop\python-work\Vendors - Monthly.xlsx'

wb = openpyxl.load_workbook(workbook_path)

ws = wb['Sheet1']


default_mail_status = messagebox.askyesno("Default mail", "Do you want to use the default envelope settings for each vendor?")


first_row_skipped = False
# List to store row numbers with a cell value
rows_with_value = []

check_in_excel = messagebox.askyesno("Check number", f"Is the first check number {ws['I1'].value}")
if check_in_excel:
    check_numbers = ws['I1'].value
else:
    check_numbers = simpledialog.askinteger("Check numbers", "What is the first check number?")

qb_window.activate()
sleep(1)


column_to_check = 'B'
for cell in ws[column_to_check]:
    # Skip the first row (header row)
    if not first_row_skipped:
        first_row_skipped = True
        continue

    if cell.value is not None:
        print(f"Cell {cell.coordinate} contains text: {cell.value}")
        rows_with_value.append(cell.coordinate[1:])

print(rows_with_value)

# # Write the bills to QB
qb_window.activate()
sleep(1)


open_pay_bill_menu()

payment_date = set_payment_date()


# Pay the bills

bills_counter = 0
for j in rows_with_value:
    print("starting main loop")
    # print(j)
    # print(rows_with_value[0])
    sleep(.5)
    # move focus to due on or before
    if j == rows_with_value[0]:
        sleep(.5)
        hotkey('shift', 'tab')
        sleep(.5)
        press('space')
        sleep(.5)
        # move focus to due date
        press('tab', presses=2)
        sleep(.1)   
        # Write end of this month
        sleep(1)
        write(payment_date)
        sleep(1)
    else:
        press('space')
        press('tab', presses=2)
    sleep(.5)
    # move focus to filter by (selects vendor)
    press('tab')
    sleep(1)
    sleep(1)
    pyautogui.write(ws['A'+j].value)
    sleep(2)
    press('tab')
    sleep(2)
    # Selects the pay all bills button
    for i in range(11):
        print(i)
        hotkey('shift', 'tab')
        sleep(.1)
    sleep(2)
    press('space')
    sleep(2)
    final_confirmation = messagebox.askyesno("Final Confirmation", "Is this all correct for this vendor? Press yes when correct, or no if you want to stop this program.")
    if not final_confirmation:
        break
    press('p')
    # press('tab')
    # sleep(2)
    # press('space')
    # sleep(1)
    press('p')
    sleep(2)
    bills_counter += 1
    if bills_counter == len(rows_with_value):
        press('tab')
        sleep(2)
        press('enter')
        sleep(2)
    else:
        press('enter')
        sleep(3)

sleep(1)
press('tab')
sleep(.2)
write(str(check_numbers))
sleep(1)
press('enter')

print_ready = messagebox.askyesno("Ready?", "Have your checks printed?")
# pyautogui.write(ws['A'+'2'].value)
# Print any checks needed

if print_ready:
    if default_mail_status == True:
        mail_col = 'G'
    else:
        mail_col = 'F'

    for row in rows_with_value:
        if ws[mail_col+row].value is not None:
            a_value = ws['A' + str(row)].value
            # Path to your Word document
            print(a_value)
        
            file_path = f"C:\\Users\\Michael\\Desktop\\CEnvelopes\\{a_value}.docx"

            # Start an instance of Word
            word = win32com.client.Dispatch('Word.Application')

            # Open the document

            doc = word.Documents.Open(file_path)

            # Print the document to the default printer
            doc.PrintOut()

            # Close the documentC
            doc.Close()

            # Quit Word
            word.Quit()
            sleep(2)
        else:
            print("skipped envelope")